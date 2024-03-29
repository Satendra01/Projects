from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

siteUrl ='https://leetcode.com/problemset/all/'
questionNameList = []
questionUrlList = []
questionDifficultyList = []
# driver=None

def xcelSheet():

    excelFileName = 'LeetCode.xlsx'
    sheetName = 'Top Interview Questions'

    df = pd.DataFrame({
        'Question Name': questionNameList,
        'Question Url': questionUrlList,
        'Question Difficulty': questionDifficultyList
    })

    wb = Workbook()
    sheet1 = wb.create_sheet(sheetName)
    sheet1.cell(1, 1, 'Question Name')
    sheet1.cell(1, 2, 'Question URL')
    sheet1.cell(1, 3, 'Question Difficulty')

    for i in range(0, df.__len__()):
        sheet1.cell(i + 2, 1, df['Question Name'][i])
        sheet1.cell(i + 2, 2, df['Question Url'][i])
        sheet1.cell(i + 2, 3, df['Question Difficulty'][i])

    wb.save(excelFileName)
    wb.close()
    print("     -----------> Excel sheet created")


def openBrowser(url):
    print("     -----------> Opening Browser")
    options = webdriver.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument('--incognito')
    # options.add_argument('--headless')

    # driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

    # headless browser
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),
                              options=options)

    # driver.get(url)
    # driver.maximize_window()
    return driver

def sign_into_leetcode():

    url = 'https://leetcode.com/accounts/login'
    driver=openBrowser(url)
    driver.get(url)
    WebDriverWait(driver, timeout=15).until_not(lambda x: x.find_element(By.CLASS_NAME,"spinner").is_displayed())
    try:
        emailInput = driver.find_element(By.ID,'id_login')
        passwordInput = driver.find_element(By.ID,'id_password')
        loginButton = driver.find_element(By.ID,'signin_btn')
        emailInput.send_keys('<username>')
        passwordInput.send_keys('<password>')
        loginButton.click()
    except:
        pass
    finally:
        WebDriverWait(driver, timeout=25).until(lambda x: x.find_element(By.CLASS_NAME,"container").is_displayed())
    return driver
def closeBrowser(driver):
    print("     -----------> Closing Browser")
    driver.close()


def fetchPageData(pageUrl):
    sleepTime = 3

    print("Page URL: ", pageUrl)
    # browser = openBrowser(pageUrl)
    driver.get(pageUrl)
    time.sleep(sleepTime)
    pageSource = driver.page_source
    WebDriverWait(driver, 10).until(EC.title_contains("Problems - LeetCode"))
    # print(f"title is: {browser.title}")

    soup = BeautifulSoup(pageSource, 'html.parser')
    if (driver.title == "Problems - LeetCode"):
        print(
            "\n\n                     ------------------- Parsing data -------------------\n\n"
        )
        newSoup = BeautifulSoup(pageSource, 'html.parser')
        questionBlock = newSoup.find('div', role='rowgroup')
        questionList = questionBlock.find_all('div', role='row')
        # print(f"Total {questionList.__len__()} data fetched ")

        for question in questionList:
            row = question.find_all('div', role='cell')
            questionName = row[1].find('a').text
            questionUrl = row[1].find('a')['href']
            questionUrl = 'https://leetcode.com' + questionUrl
            questionDifficulty = row[4].find('span').text
            if row[0].find('svg',xmlns='http://www.w3.org/2000/svg') is not None:
                questionNameList.append(questionName)
                questionUrlList.append(questionUrl)
                questionDifficultyList.append(questionDifficulty)
            
        print("     -----------> Done")
        # closeBrowser(browser)

    else:
        print("Page does not exist o connection Failed, status code: ",
              soup.status_code)
    return


def getData(driver):

    try:

        # time.sleep(2)
        # pageSource = driver.page_source
        driver.get(siteUrl)
        WebDriverWait(driver, 10).until(EC.title_contains("Problems - LeetCode"))
        # soup = BeautifulSoup(pageSource, 'html.parser')
        print("Entering page")
        if (driver.title == "Problems - LeetCode"):
            print("Inside Page")
            # Fetching total number of pages
            # totalQuestion = soup.find('div', class_="text-label-2 dark:text-dark-label-2 mr-2").find_all('span')[1]
            # totalQuestion = totalQuestion.text.split('/')[1]
            # totalQuestion = int(totalQuestion)
            # print(f"Total {totalQuestion} questions available")
            # totalPage = (totalQuestion+49)//50
            totalPage=10
            print(f"Total {totalPage} pages available")
            # closeBrowser(browser)

            # Fetching data from each page
            for page in range(1, totalPage + 1):
                print(
                    f"\n\n                     ------------------- Fetching Page {page} -------------------\n\n"
                )
                pageUrl = siteUrl + '?page=' + str(page)
                fetchPageData(pageUrl)

            print("     -----------> Done all pages ")
            print(f"Total {questionNameList.__len__()} questions fetched")
            xcelSheet()

        else:
            print("Connection Failed")
            return

    except Exception as e:
        print("Some error occured, error: ", e)
        return


if __name__ == "__main__":

    driver=sign_into_leetcode()
    getData(driver)
    driver.close()
